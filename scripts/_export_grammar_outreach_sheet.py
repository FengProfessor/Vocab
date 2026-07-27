# -*- coding: utf-8 -*-
"""Export outreach sheet for 1 grammar class (~20, 25k, T3+T5 22h)."""
import glob
import re
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

path = glob.glob(r"D:\Download\Khóa chuyên đề*.xlsx")[0]
ws_in = load_workbook(path, data_only=True).active
rows = list(ws_in.iter_rows(values_only=True))[1:]


def norm_phone(p):
    if p is None:
        return ""
    s = str(p).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    if len(s) == 9 and not s.startswith("0"):
        s = "0" + s
    if len(s) == 11 and s.startswith("84"):
        return "0" + s[2:]
    return s


def classify(nv):
    s = str(nv or "")
    if "ĐĂNG KÝ" in s:
        return "register", "Đăng ký học tiếp"
    if "app" in s.lower() or "tự học" in s:
        return "app", "Tự học app"
    if "DỪNG" in s:
        return "stop", "Dừng / khóa sau"
    return "other", s[:40]


def pkg_short(p):
    p = str(p or "").strip()
    if "Gói A" in p:
        return "A (form 20k)"
    if "Gói B" in p:
        return "B (kèm sát)"
    if "cũng được" in p or "tùy thầy" in p.lower():
        return "Flex (gói nào cũng được)"
    return p or ""


def parse_slots(sc):
    sc = str(sc or "")
    slots = []
    if "Thứ 3 + Thứ 6" in sc:
        slots.append("T3+T6 17h30")
    if "22h00" in sc or "22h" in sc:
        slots.append("T3+T5 22h")
    if "Chủ Nhật" in sc or "Chủ nhật" in sc:
        slots.append("T4+CN")
    if "19h-20h" in sc:
        slots.append("xin 19-20h")
    if "chưa biết" in sc.lower():
        slots.append("chưa biết lịch")
    return slots, sc.strip()


def fit_22h(slots, intent):
    if intent != "register":
        return ""
    if "T3+T5 22h" in slots:
        return "Có 22h"
    if "chưa biết lịch" in slots:
        return "Hỏi lại"
    if slots:
        return "Chỉ ca khác — hỏi 22h"
    return "Hỏi lại"


by = OrderedDict()
order = []
for r in rows:
    ph = norm_phone(r[2])
    key = ph or ("n:" + str(r[1]).strip().lower())
    if key not in by:
        order.append(key)
    by[key] = r

records = []
for key in order:
    r = by[key]
    intent_c, intent_l = classify(r[4])
    slots, slot_raw = parse_slots(r[6])
    prio_map = {
        "register": "P0 — Mời lớp",
        "app": "P1 — App (dự bị)",
        "stop": "P2 — Không mời",
        "other": "P2 — Khác",
    }
    prio = prio_map[intent_c]
    if intent_c == "register":
        if "T3+T5 22h" in slots:
            suggest = "Tin chốt: 1 lớp ~20 · 25k · T3-T5 22h — ask OK + CK"
        elif "Chỉ ca khác" in fit_22h(slots, intent_c) or fit_22h(slots, intent_c) == "Chỉ ca khác — hỏi 22h":
            suggest = "Hỏi: vào được ca 22h T3-T5 không? (1 lớp duy nhất)"
        else:
            suggest = "Hỏi lịch + chốt 25k / 22h"
    elif intent_c == "app":
        suggest = "Chưa mời lớp. Chỉ khi thiếu slot: offer 1 lớp 25k 22h"
    else:
        suggest = "Không chase. Có thể 1 tin cảm ơn + app (tuỳ)"

    records.append(
        {
            "prio": prio,
            "intent_c": intent_c,
            "name": str(r[1] or "").strip(),
            "phone": norm_phone(r[2]),
            "grade": str(r[3] or "").strip(),
            "intent": intent_l,
            "pkg": pkg_short(r[5]) if intent_c == "register" else "",
            "slots": ", ".join(slots) if slots else "",
            "fit22": fit_22h(slots, intent_c),
            "pain": str(r[7] or "").strip() if intent_c == "register" else "",
            "note": str(r[8] or "").strip() if r[8] else "",
            "ts": str(r[0])[:19] if r[0] else "",
            "suggest": suggest,
        }
    )

prio_rank = {
    "P0 — Mời lớp": 0,
    "P1 — App (dự bị)": 1,
    "P2 — Không mời": 2,
    "P2 — Khác": 3,
}
fit_rank = {"Có 22h": 0, "Hỏi lại": 1, "Chỉ ca khác — hỏi 22h": 2, "": 3}
records.sort(
    key=lambda x: (
        prio_rank.get(x["prio"], 9),
        fit_rank.get(x["fit22"], 9),
        x["grade"],
        x["name"],
    )
)

wb = Workbook()
ws = wb.active
ws.title = "Nhắn tin lớp"

headers = [
    "Ưu tiên",
    "STT",
    "Họ tên",
    "SĐT Zalo",
    "Lớp",
    "Nguyện vọng form",
    "Gói form",
    "Ca form (rút gọn)",
    "Fit ca 22h?",
    "Pain / khó khăn",
    "Lời nhắn form",
    "Gợi ý tin nhắn",
    "Đã nhắn?",
    "OK 25k?",
    "OK ca 22h?",
    "Đã CK?",
    "Vào lớp?",
    "Ghi chú thầy",
    "Submit form",
]

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
fills = {
    "P0 — Mời lớp": PatternFill("solid", fgColor="C6EFCE"),
    "P1 — App (dự bị)": PatternFill("solid", fgColor="FFEB9C"),
    "P2 — Không mời": PatternFill("solid", fgColor="F2F2F2"),
    "P2 — Khác": PatternFill("solid", fgColor="F2F2F2"),
}
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
wrap = Alignment(wrap_text=True, vertical="center")

for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin

for i, x in enumerate(records, 1):
    vals = [
        x["prio"],
        i,
        x["name"],
        x["phone"],
        x["grade"],
        x["intent"],
        x["pkg"],
        x["slots"],
        x["fit22"],
        x["pain"],
        x["note"],
        x["suggest"],
        "",
        "",
        "",
        "",
        "",
        "",
        x["ts"],
    ]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(i + 1, col, v)
        cell.border = thin
        cell.alignment = wrap
        if col == 1:
            cell.fill = fills.get(x["prio"], PatternFill())
        if col == 9 and x["fit22"] == "Có 22h":
            cell.fill = PatternFill("solid", fgColor="A9D08E")
        if col == 9 and x["fit22"] == "Chỉ ca khác — hỏi 22h":
            cell.fill = PatternFill("solid", fgColor="FCE4D6")

for col_letter, opts in [
    ("M", '"Chưa,Đã nhắn,Đã reply"'),
    ("N", '"Chưa hỏi,OK 25k,Không OK"'),
    ("O", '"Chưa hỏi,OK 22h,Không OK"'),
    ("P", '"Chưa,Đã CK,Chờ CK"'),
    ("Q", '"Chưa,Có,Waitlist,Out"'),
]:
    dv = DataValidation(type="list", formula1=opts, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}200")

widths = [16, 5, 22, 14, 12, 16, 18, 22, 18, 36, 28, 40, 12, 12, 12, 10, 12, 22, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 32
ws.auto_filter.ref = f"A1:S{len(records) + 1}"
ws.freeze_panes = "C2"

# Summary
ws2 = wb.create_sheet("Tóm tắt & tin mẫu", 0)
ws2["A1"] = (
    "LỚP CHUYÊN ĐỀ NGỮ PHÁP — 1 lớp · ~20 HV · 25k/buổi · T3+T5 22h00–23h30"
)
ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")
ws2.merge_cells("A1:B1")

n_reg = sum(1 for x in records if x["intent_c"] == "register")
n_22 = sum(1 for x in records if x["fit22"] == "Có 22h")
n_app = sum(1 for x in records if x["intent_c"] == "app")
n_stop = sum(1 for x in records if x["intent_c"] == "stop")

summary = [
    ("", ""),
    ("Tổng unique form", len(records)),
    ("P0 — Mời lớp (đăng ký form)", n_reg),
    ("  · Đã chọn ca 22h", n_22),
    ("  · Chỉ ca khác / hỏi lại", n_reg - n_22),
    ("P1 — App (dự bị đủ 20)", n_app),
    ("P2 — Dừng / không mời", n_stop),
    ("", ""),
    ("Giá chốt", "25.000đ/buổi"),
    ("Ca chốt", "Thứ 3 + Thứ 5, 22h00 – 23h30"),
    ("Sĩ số", "Target 20 (mở lớp ≥15 CK)"),
    ("", ""),
    ("CÁCH DÙNG SHEET", ""),
    ("1", "Mở tab «P0 — Mời lớp (nhanh)» — nhắn hết list này trước"),
    ("2", "Tab «Nhắn tin lớp» = full 58 người + filter Ưu tiên"),
    ("3", "Cột Đã nhắn / OK 25k / OK 22h / CK / Vào lớp — tick khi làm"),
    ("4", "Thiếu slot → P1, tối đa 4–5 bạn L11–12"),
    ("5", "Không mở ca 2, không giảm dưới 25k"),
]
for i, (a, b) in enumerate(summary, 3):
    ws2.cell(i, 1, a)
    ws2.cell(i, 2, b)
    if a and b == "" and not str(a).startswith(" "):
        ws2.cell(i, 1).font = Font(bold=True)

ws2["A23"] = "TIN MẪU — Broadcast P0"
ws2["A23"].font = Font(bold=True, size=12, color="C00000")
ws2["A24"] = (
    "Chốt lớp chuyên đề ngữ pháp (1 lớp duy nhất):\n\n"
    "• Sĩ số: ~20 bạn\n"
    "• Lịch: Thứ 3 + Thứ 5, 22h00 – 23h30\n"
    "• Học phí: 25.000đ/buổi\n"
    "• Không mở ca/gói khác kỳ này\n\n"
    'Em/gia đình OK lịch + học phí → reply: "ĐĂNG KÝ + Họ tên"\n'
    "Deadline: ____ (điền ngày giờ)\n"
    "Đủ 20 hoặc hết hạn là khóa danh sách."
)
ws2["A24"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.merge_cells("A24:B31")
ws2.row_dimensions[24].height = 150

ws2["A33"] = "TIN MẪU — Riêng (chỉ chọn ca 17h30 / ca khác)"
ws2["A33"].font = Font(bold=True, size=12)
ws2["A34"] = (
    "Em ơi kỳ này thầy chỉ mở 1 lớp: T3+T5 22h–23h30, 25k/buổi (~20 bạn).\n"
    "Form em chọn ca khác — em vào được ca 22h không?\n"
    'OK thì reply "ĐĂNG KÝ", không kịp ca thì mình ghi danh khóa sau.'
)
ws2["A34"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.merge_cells("A34:B37")
ws2.row_dimensions[34].height = 70

ws2["A39"] = "TIN MẪU — P1 app khi thiếu slot"
ws2["A39"].font = Font(bold=True, size=12)
ws2["A40"] = (
    "Em từng học free với thầy — còn 1–2 slot lớp ngữ pháp:\n"
    "T3+T5 22h, 25k/buổi, ~20 bạn. Em có muốn vào không?\n"
    "OK reply nhanh giúp thầy."
)
ws2["A40"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.merge_cells("A40:B43")
ws2.row_dimensions[40].height = 60

ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 55

# P0 quick
ws3 = wb.create_sheet("P0 — Mời lớp (nhanh)", 1)
h2 = [
    "STT",
    "Họ tên",
    "SĐT Zalo",
    "Lớp",
    "Gói form",
    "Fit 22h?",
    "Pain / khó khăn",
    "Lời nhắn form",
    "Gợi ý tin",
    "Đã nhắn?",
    "OK 25k?",
    "OK 22h?",
    "Đã CK?",
    "Vào lớp?",
    "Ghi chú",
]
for col, h in enumerate(h2, 1):
    c = ws3.cell(1, col, h)
    c.fill = header_fill
    c.font = header_font
    c.border = thin
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

p0 = [x for x in records if x["intent_c"] == "register"]
for i, x in enumerate(p0, 1):
    vals = [
        i,
        x["name"],
        x["phone"],
        x["grade"],
        x["pkg"],
        x["fit22"],
        x["pain"],
        x["note"],
        x["suggest"],
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    for col, v in enumerate(vals, 1):
        cell = ws3.cell(i + 1, col, v)
        cell.border = thin
        cell.alignment = wrap
        if col == 6 and x["fit22"] == "Có 22h":
            cell.fill = PatternFill("solid", fgColor="A9D08E")
        if col == 6 and x["fit22"] == "Chỉ ca khác — hỏi 22h":
            cell.fill = PatternFill("solid", fgColor="FCE4D6")

for col, w in enumerate([5, 22, 14, 12, 18, 18, 32, 28, 38, 12, 12, 12, 10, 12, 18], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.row_dimensions[1].height = 30
ws3.freeze_panes = "B2"
ws3.auto_filter.ref = f"A1:O{len(p0) + 1}"

for col_letter, opts in [
    ("J", '"Chưa,Đã nhắn,Đã reply"'),
    ("K", '"Chưa hỏi,OK 25k,Không OK"'),
    ("L", '"Chưa hỏi,OK 22h,Không OK"'),
    ("M", '"Chưa,Đã CK,Chờ CK"'),
    ("N", '"Chưa,Có,Waitlist,Out"'),
]:
    dv = DataValidation(type="list", formula1=opts, allow_blank=True)
    ws3.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}50")

out = r"D:\Download\Lop-ngu-phap-nhan-tin-1-lop-25k.xlsx"
wb.save(out)
print("SAVED", out)
print("ALL", len(records), "P0", len(p0), "P0_22h", n_22)
